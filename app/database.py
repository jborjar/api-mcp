import pyodbc
import httpx
import smtplib
import json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.utils import make_msgid
from datetime import datetime
from hdbcli import dbapi
from config import get_settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


def get_mssql_connection(database: str | None = None):
    """Obtiene conexión a MSSQL. Si database es None, usa la configurada en settings."""
    settings = get_settings()
    db = database if database else settings.MSSQL_DATABASE
    connection_string = (
        f"DRIVER={{ODBC Driver 18 for SQL Server}};"
        f"SERVER={settings.MSSQL_HOST},{settings.MSSQL_PORT};"
        f"DATABASE={db};"
        f"UID={settings.MSSQL_USER};"
        f"PWD={settings.MSSQL_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    return pyodbc.connect(connection_string)


def drop_and_create_database() -> bool:
    """
    Elimina la base de datos si existe y la recrea desde cero.
    Retorna True si se creó exitosamente.
    """
    settings = get_settings()

    # Conectar a master para eliminar/crear la base de datos
    conn = get_mssql_connection(database="master")
    conn.autocommit = True
    cursor = conn.cursor()

    try:
        # Eliminar base de datos si existe
        cursor.execute(f"""
            IF EXISTS (SELECT name FROM sys.databases WHERE name = '{settings.MSSQL_DATABASE}')
            BEGIN
                ALTER DATABASE [{settings.MSSQL_DATABASE}] SET SINGLE_USER WITH ROLLBACK IMMEDIATE;
                DROP DATABASE [{settings.MSSQL_DATABASE}];
            END
        """)

        # Crear base de datos nueva
        cursor.execute(f"CREATE DATABASE [{settings.MSSQL_DATABASE}]")

        return True
    finally:
        cursor.close()
        conn.close()


def ensure_database_exists() -> bool:
    """
    Verifica si la base de datos existe en MSSQL, si no existe la crea.
    Retorna True si ya existía o fue creada exitosamente.
    """
    settings = get_settings()

    # Conectar a master para verificar/crear la base de datos
    conn = get_mssql_connection(database="master")
    conn.autocommit = True
    cursor = conn.cursor()

    try:
        cursor.execute(
            "SELECT COUNT(*) FROM sys.databases WHERE name = ?",
            (settings.MSSQL_DATABASE,)
        )
        exists = cursor.fetchone()[0] > 0

        if not exists:
            cursor.execute(f"CREATE DATABASE [{settings.MSSQL_DATABASE}]")

        return True
    finally:
        cursor.close()
        conn.close()


def ensure_table_sap_empresas_exists() -> bool:
    """
    Verifica si la tabla SAP_EMPRESAS existe, si no existe la crea.
    Retorna True si ya existía o fue creada exitosamente.
    """
    conn = get_mssql_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = 'SAP_EMPRESAS'
        """)
        exists = cursor.fetchone()[0] > 0

        if not exists:
            cursor.execute("""
                CREATE TABLE SAP_EMPRESAS (
                    Instancia NVARCHAR(100) NOT NULL,
                    SL BIT NOT NULL DEFAULT 0,
                    Prueba BIT NOT NULL DEFAULT 0,
                    SLP BIT NOT NULL DEFAULT 0,
                    PrintHeadr NVARCHAR(255),
                    CompnyAddr NVARCHAR(500),
                    TaxIdNum NVARCHAR(50),
                    PRIMARY KEY (Instancia)
                )
            """)
            conn.commit()

        return True
    finally:
        cursor.close()
        conn.close()


def ensure_table_sap_proveedores_exists() -> bool:
    """
    Verifica si la tabla SAP_PROVEEDORES existe, si no existe la crea.
    Retorna True si ya existía o fue creada exitosamente.
    """
    conn = get_mssql_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = 'SAP_PROVEEDORES'
        """)
        exists = cursor.fetchone()[0] > 0

        if not exists:
            cursor.execute("""
                CREATE TABLE SAP_PROVEEDORES (
                    -- Identificación básica
                    Instancia NVARCHAR(100) NOT NULL,
                    CardCode NVARCHAR(50) NOT NULL,
                    CardName NVARCHAR(200),
                    GroupCode INT,
                    FederalTaxID NVARCHAR(50),
                    -- Fechas de auditoría
                    CreateDate DATE,
                    CreateTime TIME,
                    UpdateDate DATE,
                    UpdateTime TIME,
                    -- Dirección principal
                    Address NVARCHAR(500),
                    Block NVARCHAR(200),
                    ZipCode NVARCHAR(20),
                    City NVARCHAR(100),
                    County NVARCHAR(100),
                    BillToState NVARCHAR(10),
                    Country NVARCHAR(10),
                    -- Dirección postal / Envío
                    MailAddress NVARCHAR(500),
                    MailZipCode NVARCHAR(20),
                    ShipToState NVARCHAR(10),
                    ShipToDefault NVARCHAR(100),
                    -- Contacto
                    Phone1 NVARCHAR(50),
                    Phone2 NVARCHAR(50),
                    Fax NVARCHAR(50),
                    Cellular NVARCHAR(50),
                    EmailAddress NVARCHAR(200),
                    ContactPerson NVARCHAR(200),
                    -- Condiciones financieras
                    PayTermsGrpCode INT,
                    PeymentMethodCode NVARCHAR(50),
                    CreditLimit DECIMAL(18,2),
                    MaxCommitment DECIMAL(18,2),
                    DiscountPercent DECIMAL(5,2),
                    PriceListNum INT,
                    Currency NVARCHAR(10),
                    -- Impuestos y deducciones
                    DeductibleAtSource NVARCHAR(10),
                    DeductionPercent DECIMAL(5,2),
                    DeductionValidUntil DATE,
                    VatGroupLatinAmerica NVARCHAR(20),
                    -- Datos bancarios
                    DefaultBankCode NVARCHAR(50),
                    DefaultAccount NVARCHAR(100),
                    BankCountry NVARCHAR(10),
                    HouseBank NVARCHAR(50),
                    HouseBankCountry NVARCHAR(10),
                    HouseBankAccount NVARCHAR(50),
                    HouseBankBranch NVARCHAR(20),
                    HouseBankIBAN NVARCHAR(50),
                    IBAN NVARCHAR(50),
                    CreditCardCode INT,
                    CreditCardNum NVARCHAR(50),
                    CreditCardExpiration DATE,
                    DebitorAccount NVARCHAR(50),
                    -- Saldos y oportunidades
                    CurrentAccountBalance DECIMAL(18,2),
                    OpenDeliveryNotesBalance DECIMAL(18,2),
                    OpenOrdersBalance DECIMAL(18,2),
                    OpenChecksBalance DECIMAL(18,2),
                    OpenOpportunities INT,
                    -- Estado del proveedor
                    Valid NVARCHAR(10),
                    Frozen NVARCHAR(10),
                    BlockDunning NVARCHAR(10),
                    BackOrder NVARCHAR(10),
                    PartialDelivery NVARCHAR(10),
                    PRIMARY KEY (Instancia, CardCode)
                )
            """)
            conn.commit()

        return True
    finally:
        cursor.close()
        conn.close()


def create_or_update_vista_maestro_proveedores() -> bool:
    """
    Crea o actualiza la vista maestro_proveedores.
    La vista pivotea los proveedores mostrando CardName, GroupCode, FederalTaxID
    y una columna por cada instancia con el CardCode correspondiente.
    """
    conn = get_mssql_connection()
    cursor = conn.cursor()

    try:
        # Obtener las instancias únicas de SAP_PROVEEDORES
        cursor.execute("SELECT DISTINCT Instancia FROM SAP_PROVEEDORES ORDER BY Instancia")
        instancias = [row[0] for row in cursor.fetchall()]

        if not instancias:
            return False

        # Construir columnas para el PIVOT
        pivot_columns = ", ".join([f"[{inst}]" for inst in instancias])

        # Crear o actualizar la vista
        view_sql = f"""
            CREATE OR ALTER VIEW maestro_proveedores AS
            SELECT
                CardName,
                GroupCode,
                FederalTaxID,
                {pivot_columns}
            FROM (
                SELECT CardName, GroupCode, FederalTaxID, Instancia, CardCode
                FROM SAP_PROVEEDORES
            ) AS SourceTable
            PIVOT (
                MAX(CardCode)
                FOR Instancia IN ({pivot_columns})
            ) AS PivotTable
        """

        cursor.execute(view_sql)
        conn.commit()
        return True
    finally:
        cursor.close()
        conn.close()


def get_maestro_proveedores(
    top: int | None = None,
    card_name: str | None = None,
    federal_tax_id: str | None = None
) -> dict:
    """
    Consulta la vista maestro_proveedores.
    Retorna proveedores con CardName, GroupCode, FederalTaxID y CardCode por instancia.
    """
    conn = get_mssql_connection()
    cursor = conn.cursor()

    try:
        # Verificar si la vista existe
        cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.VIEWS
            WHERE TABLE_NAME = 'maestro_proveedores'
        """)
        if cursor.fetchone()[0] == 0:
            return {"success": False, "error": "Vista maestro_proveedores no existe. Ejecute /inicializa_datos primero."}

        # Construir query con filtros
        query = "SELECT * FROM maestro_proveedores WHERE 1=1"
        params = []

        if card_name:
            query += " AND CardName LIKE ?"
            params.append(f"%{card_name}%")

        if federal_tax_id:
            query += " AND FederalTaxID LIKE ?"
            params.append(f"%{federal_tax_id}%")

        # Ordenar por CardName
        query += " ORDER BY CardName"

        # Aplicar TOP si se especifica
        if top:
            query = query.replace("SELECT *", f"SELECT TOP {top} *")

        cursor.execute(query, params)
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()

        # Convertir a lista de diccionarios
        proveedores = []
        for row in rows:
            prov = {}
            for i, col in enumerate(columns):
                prov[col] = row[i]
            proveedores.append(prov)

        return {
            "success": True,
            "total": len(proveedores),
            "columnas": columns,
            "proveedores": proveedores
        }
    finally:
        cursor.close()
        conn.close()


def get_hana_connection():
    settings = get_settings()
    return dbapi.connect(
        address=settings.SAP_HANA_HOST,
        port=settings.SAP_HANA_PORT,
        user=settings.SAP_HANA_USER,
        password=settings.SAP_HANA_PASSWORD
    )


def get_empresas_sap() -> list[str]:
    conn = get_hana_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT SCHEMA_NAME
        FROM SCHEMAS
        WHERE SCHEMA_NAME NOT LIKE 'B1%'
          AND SCHEMA_NAME NOT LIKE '_SYS%'
          AND SCHEMA_NAME NOT LIKE 'SAP%'
          AND SCHEMA_NAME NOT LIKE 'XSSQLCC%'
          AND SCHEMA_NAME NOT LIKE '%_PRUEBAS%'
          AND SCHEMA_NAME NOT LIKE '%_MIGRACION%'
          AND SCHEMA_NAME NOT IN ('SYS', 'SYSTEM', 'SBOCOMMON', 'SLDDATA', 'HANA_XS_BASE', 'UIS', 'COMMON')
        ORDER BY SCHEMA_NAME
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [row[0] for row in rows]


def schema_exists_in_hana(schema_name: str) -> bool:
    """Verifica si un schema existe en HANA."""
    conn = get_hana_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(*) FROM SCHEMAS WHERE SCHEMA_NAME = ?",
        (schema_name,)
    )
    result = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    return result > 0


def get_oadm_data(schema_name: str) -> dict:
    """Obtiene PrintHeadr, CompnyAddr, TaxIdNum de la tabla OADM de una instancia."""
    conn = get_hana_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f'''
            SELECT "PrintHeadr", "CompnyAddr", "TaxIdNum"
            FROM "{schema_name}"."OADM"
            LIMIT 1
        ''')
        row = cursor.fetchone()
        if row:
            return {
                "PrintHeadr": row[0] or "",
                "CompnyAddr": row[1] or "",
                "TaxIdNum": row[2] or ""
            }
    except Exception:
        pass
    finally:
        cursor.close()
        conn.close()
    return {"PrintHeadr": "", "CompnyAddr": "", "TaxIdNum": ""}


def inicializa_sap_empresas() -> dict:
    """
    Inicializa la tabla SAP_EMPRESAS:
    1. Elimina y recrea la base de datos desde cero
    2. Crea la tabla SAP_EMPRESAS
    3. Crea la tabla USER_SESSIONS (para el sistema de sesiones)
    4. Obtiene las instancias de HANA
    5. Para cada instancia, verifica si existe versión _PRUEBAS
    6. Obtiene datos de OADM
    7. Inserta en SAP_EMPRESAS
    """
    from session import ensure_sessions_table_exists

    # Eliminar y recrear la base de datos
    drop_and_create_database()
    ensure_table_sap_empresas_exists()

    # Recrear tabla de sesiones (necesaria para el sistema de autenticación)
    ensure_sessions_table_exists()

    empresas = get_empresas_sap()

    mssql_conn = get_mssql_connection()
    mssql_cursor = mssql_conn.cursor()

    insertados = 0
    errores = []

    for instancia in empresas:
        try:
            # Verificar si existe versión _PRUEBAS
            tiene_pruebas = schema_exists_in_hana(f"{instancia}_PRUEBAS")

            # Obtener datos de OADM
            oadm = get_oadm_data(instancia)

            # Si PrintHeadr está vacío, usar el nombre de la instancia
            print_headr = oadm["PrintHeadr"] if oadm["PrintHeadr"] else instancia

            # Insertar en SAP_EMPRESAS
            mssql_cursor.execute(
                """
                INSERT INTO SAP_EMPRESAS (Instancia, Prueba, PrintHeadr, CompnyAddr, TaxIdNum)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    instancia,
                    1 if tiene_pruebas else 0,
                    print_headr,
                    oadm["CompnyAddr"],
                    oadm["TaxIdNum"]
                )
            )
            insertados += 1
        except Exception as e:
            errores.append({"instancia": instancia, "error": str(e)})

    mssql_conn.commit()
    mssql_cursor.close()
    mssql_conn.close()

    return {
        "total_empresas": len(empresas),
        "insertados": insertados,
        "errores": errores
    }


def test_service_layer_login(company_db: str) -> dict:
    """
    Prueba login en SAP B1 Service Layer para una instancia específica.
    Retorna el resultado del intento de conexión.
    """
    settings = get_settings()

    if not settings.SAP_B1_SERVICE_LAYER_URL:
        return {"success": False, "error": "SAP_B1_SERVICE_LAYER_URL no configurada"}

    login_url = f"{settings.SAP_B1_SERVICE_LAYER_URL}Login"

    payload = {
        "CompanyDB": company_db,
        "UserName": settings.SAP_B1_USER,
        "Password": settings.SAP_B1_PASSWORD
    }

    try:
        with httpx.Client(verify=False, timeout=30.0) as client:
            response = client.post(login_url, json=payload)

            if response.status_code == 200:
                # Hacer logout para liberar la sesión
                session_id = response.cookies.get("B1SESSION")
                if session_id:
                    logout_url = f"{settings.SAP_B1_SERVICE_LAYER_URL}Logout"
                    client.post(logout_url, cookies={"B1SESSION": session_id})
                return {"success": True, "status_code": response.status_code}
            else:
                error_msg = response.text
                try:
                    error_json = response.json()
                    error_msg = error_json.get("error", {}).get("message", {}).get("value", response.text)
                except:
                    pass
                return {"success": False, "status_code": response.status_code, "error": error_msg}
    except httpx.TimeoutException:
        return {"success": False, "error": "Timeout de conexión"}
    except httpx.ConnectError as e:
        return {"success": False, "error": f"Error de conexión: {str(e)}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def send_email(to_email: str, subject: str, body: str, attachment: dict | None = None) -> dict:
    """
    Envía un correo electrónico usando el servidor SMTP configurado.
    attachment: dict con keys 'filename' y 'content' (string JSON)
    """
    settings = get_settings()

    if not to_email:
        return {"success": False, "error": "No hay destinatario configurado"}

    msg = MIMEMultipart()
    msg["From"] = settings.EMAIL_FROM
    msg["To"] = to_email
    msg["Subject"] = subject
    msg["Message-ID"] = make_msgid(domain="progex.grupoexpansion")
    msg["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
    msg["X-Mailer"] = "API-MCP/1.0"

    msg.attach(MIMEText(body, "plain"))

    if attachment:
        # Manejar tanto strings (JSON) como bytes (Excel)
        content = attachment["content"]
        if isinstance(content, str):
            content = content.encode("utf-8")
        part = MIMEApplication(content, Name=attachment["filename"])
        part["Content-Disposition"] = f'attachment; filename="{attachment["filename"]}"'
        msg.attach(part)

    try:
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=30) as server:
            server.sendmail(msg["From"], to_email, msg.as_string())
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


def update_service_layer_status(instancia: str, status: bool) -> None:
    """
    Actualiza el campo SL en SAP_EMPRESAS para una instancia.
    """
    mssql_conn = get_mssql_connection()
    mssql_cursor = mssql_conn.cursor()
    mssql_cursor.execute(
        "UPDATE SAP_EMPRESAS SET SL = ? WHERE Instancia = ?",
        (1 if status else 0, instancia)
    )
    mssql_conn.commit()
    mssql_cursor.close()
    mssql_conn.close()


def get_proveedores_sl(
    company_db: str,
    top: int | None = None,
    card_code: str | None = None,
    card_name: str | None = None,
    federal_tax_id: str | None = None
) -> dict:
    """
    Obtiene los proveedores (BusinessPartners con CardType='S') desde Service Layer.
    top: limita el número de registros retornados (opcional)
    card_code: filtra por CardCode específico (opcional)
    card_name: filtra por CardName que contenga el valor (opcional)
    federal_tax_id: filtra por FederalTaxID que contenga el valor (opcional)
    """
    settings = get_settings()

    if not settings.SAP_B1_SERVICE_LAYER_URL:
        return {"success": False, "error": "SAP_B1_SERVICE_LAYER_URL no configurada"}

    login_url = f"{settings.SAP_B1_SERVICE_LAYER_URL}Login"

    payload = {
        "CompanyDB": company_db,
        "UserName": settings.SAP_B1_USER,
        "Password": settings.SAP_B1_PASSWORD
    }

    try:
        with httpx.Client(verify=False, timeout=60.0) as client:
            # Login
            response = client.post(login_url, json=payload)

            if response.status_code != 200:
                error_msg = response.text
                try:
                    error_json = response.json()
                    error_msg = error_json.get("error", {}).get("message", {}).get("value", response.text)
                except:
                    pass
                return {"success": False, "error": error_msg}

            session_id = response.cookies.get("B1SESSION")

            # Consultar proveedores - Campos ordenados lógicamente:
            # 1. Identificación, 2. Fechas, 3. Dirección, 4. Dirección postal, 5. Contacto,
            # 6. Financiero, 7. Impuestos, 8. Bancario, 9. Saldos, 10. Estado
            select_fields = (
                # Identificación básica
                "CardCode,CardName,GroupCode,FederalTaxID,"
                # Fechas de auditoría
                "CreateDate,CreateTime,UpdateDate,UpdateTime,"
                # Dirección principal (Calle, Colonia, CP, Ciudad, Municipio/Condado, Estado, País)
                "Address,Block,ZipCode,City,County,BillToState,Country,"
                # Dirección postal / Envío
                "MailAddress,MailZipCode,ShipToState,ShipToDefault,"
                # Contacto (Teléfonos, Fax, Email, Celular, Persona de contacto)
                "Phone1,Phone2,Fax,Cellular,EmailAddress,ContactPerson,"
                # Condiciones financieras
                "PayTermsGrpCode,PeymentMethodCode,CreditLimit,MaxCommitment,"
                "DiscountPercent,PriceListNum,Currency,"
                # Impuestos y deducciones
                "DeductibleAtSource,DeductionPercent,DeductionValidUntil,VatGroupLatinAmerica,"
                # Datos bancarios
                "DefaultBankCode,DefaultAccount,BankCountry,HouseBank,HouseBankCountry,"
                "HouseBankAccount,HouseBankBranch,HouseBankIBAN,IBAN,"
                "CreditCardCode,CreditCardNum,CreditCardExpiration,DebitorAccount,"
                # Saldos y oportunidades
                "CurrentAccountBalance,OpenDeliveryNotesBalance,OpenOrdersBalance,"
                "OpenChecksBalance,OpenOpportunities,"
                # Estado del proveedor
                "Valid,Frozen,BlockDunning,BackOrder,PartialDelivery"
            )

            # Construir filtro
            filters = ["CardType eq 'S'"]
            if card_code:
                filters.append(f"contains(CardCode, '{card_code}')")
            if card_name:
                filters.append(f"contains(CardName, '{card_name}')")
            if federal_tax_id:
                filters.append(f"contains(FederalTaxID, '{federal_tax_id}')")
            filter_str = " and ".join(filters)

            endpoint = (
                f"{settings.SAP_B1_SERVICE_LAYER_URL}BusinessPartners"
                f"?$filter={filter_str}"
                f"&$select={select_fields}"
                f"&$inlinecount=allpages"
            )

            if top is not None:
                endpoint += f"&$top={top}"

            # Usar top para maxpagesize si está presente, sino 0 (sin límite)
            max_pagesize = top if top is not None else 0
            headers = {"Prefer": f"odata.maxpagesize={max_pagesize}"}
            bp_response = client.get(
                endpoint,
                cookies={"B1SESSION": session_id},
                headers=headers
            )

            # Logout
            logout_url = f"{settings.SAP_B1_SERVICE_LAYER_URL}Logout"
            client.post(logout_url, cookies={"B1SESSION": session_id})

            if bp_response.status_code != 200:
                return {"success": False, "error": f"Error al consultar proveedores: {bp_response.text}"}

            data = bp_response.json()
            return {
                "success": True,
                "total": data.get("odata.count", len(data.get("value", []))),
                "proveedores": data.get("value", [])
            }

    except httpx.TimeoutException:
        return {"success": False, "error": "Timeout de conexión"}
    except httpx.ConnectError as e:
        return {"success": False, "error": f"Error de conexión: {str(e)}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def test_service_layer_all_instances(sap_empresas_result: dict | None = None, skip_email: bool = False) -> dict:
    """
    Prueba la conexión a Service Layer para todas las instancias de SAP.
    Actualiza el campo SL en SAP_EMPRESAS y opcionalmente envía correo con resultados.
    sap_empresas_result: resultado de inicializa_sap_empresas() para incluir en el correo
    skip_email: si es True, no envía correo (para cuando se llama desde inicializa_datos)
    """
    settings = get_settings()
    empresas = get_empresas_sap()

    resultados = {
        "exitosos": [],
        "fallidos": []
    }

    for instancia in empresas:
        resultado = test_service_layer_login(instancia)
        if resultado["success"]:
            resultados["exitosos"].append(instancia)
            update_service_layer_status(instancia, True)
        else:
            resultados["fallidos"].append({
                "instancia": instancia,
                "error": resultado.get("error", "Error desconocido")
            })
            update_service_layer_status(instancia, False)

    resultado_final = {
        "total": len(empresas),
        "exitosos": len(resultados["exitosos"]),
        "fallidos": len(resultados["fallidos"]),
        "detalle_exitosos": resultados["exitosos"],
        "detalle_fallidos": resultados["fallidos"]
    }

    # Enviar correo con resultados (solo si no se omite)
    if settings.EMAIL_SUPERVISOR and not skip_email:
        email_result = enviar_correo_inicializacion(sap_empresas_result, resultado_final, None)
        resultado_final["email_enviado"] = email_result

    return resultado_final


def enviar_correo_inicializacion(
    sap_empresas_result: dict | None,
    service_layer_result: dict | None,
    sap_proveedores_result: dict | None
) -> dict:
    """
    Envía correo con los resultados de la inicialización de datos.
    """
    settings = get_settings()

    if not settings.EMAIL_SUPERVISOR:
        return {"success": False, "error": "No hay destinatario configurado"}

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    subject = f"Inicializacion de datos - {fecha}"

    # Construir lista de fallidos para el mensaje
    fallidos_nombres = []
    if service_layer_result:
        fallidos_nombres = [f["instancia"] for f in service_layer_result.get("detalle_fallidos", [])]
    fallidos_str = ", ".join(fallidos_nombres) if fallidos_nombres else "ninguno"

    # Construir cuerpo del mensaje
    body_lines = [
        "Inicializacion de datos realizada",
        ""
    ]

    if sap_empresas_result:
        body_lines.append(f"SAP Empresas: {sap_empresas_result.get('insertados', 0)} insertadas")

    if service_layer_result:
        body_lines.append(f"Service Layer: {service_layer_result.get('exitosos', 0)} exitosos, {service_layer_result.get('fallidos', 0)} fallidos ({fallidos_str})")

    if sap_proveedores_result:
        total_actualizados = sap_proveedores_result.get('proveedores_actualizados', 0)
        total_insertados = sap_proveedores_result.get('proveedores_insertados', 0)
        total_eliminados = sap_proveedores_result.get('proveedores_eliminados', 0)
        total_instancias = sap_proveedores_result.get('total_instancias', 0)
        total_proveedores = total_actualizados + total_insertados

        body_lines.append(f"SAP Proveedores: {total_proveedores:,} sincronizados de {total_instancias} instancias")
        body_lines.append(f"  - Actualizados: {total_actualizados:,}")
        body_lines.append(f"  - Insertados: {total_insertados:,}")
        if total_eliminados > 0:
            body_lines.append(f"  - Eliminados: {total_eliminados:,}")

        # Agregar desglose por instancia (ordenado de mayor a menor)
        instancias_procesadas = sap_proveedores_result.get('instancias_procesadas', [])
        if instancias_procesadas:
            # Ordenar por cantidad de proveedores descendente
            instancias_ordenadas = sorted(instancias_procesadas, key=lambda x: x['proveedores'], reverse=True)
            body_lines.append("")
            body_lines.append("Desglose de proveedores por instancia:")
            for inst in instancias_ordenadas:
                body_lines.append(f"  {inst['instancia']}: {inst['proveedores']:,}")

    body = "\n".join(body_lines)

    # Adjuntar JSON con resumen (sin desglose detallado de proveedores)
    full_result = {
        "sap_empresas": sap_empresas_result,
        "service_layer": service_layer_result
    }
    if sap_proveedores_result:
        full_result["sap_proveedores"] = {
            "total_instancias": sap_proveedores_result.get('total_instancias', 0),
            "proveedores_actualizados": sap_proveedores_result.get('proveedores_actualizados', 0),
            "proveedores_insertados": sap_proveedores_result.get('proveedores_insertados', 0),
            "proveedores_eliminados": sap_proveedores_result.get('proveedores_eliminados', 0),
            "errores": sap_proveedores_result.get('errores', [])
        }

    attachment = {
        "filename": f"inicializacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        "content": json.dumps(full_result, indent=2, ensure_ascii=False)
    }

    return send_email(settings.EMAIL_SUPERVISOR, subject, body, attachment)


def get_instancias_con_service_layer() -> list[str]:
    """
    Obtiene las instancias de SAP_EMPRESAS que tienen SL = 1.
    En modo pruebas, solo retorna instancias que también tienen Prueba = 1.
    """
    from config import get_modo_pruebas

    conn = get_mssql_connection()
    cursor = conn.cursor()
    try:
        if get_modo_pruebas():
            cursor.execute("SELECT Instancia FROM SAP_EMPRESAS WHERE SL = 1 AND Prueba = 1")
        else:
            cursor.execute("SELECT Instancia FROM SAP_EMPRESAS WHERE SL = 1")
        rows = cursor.fetchall()
        return [row[0] for row in rows]
    finally:
        cursor.close()
        conn.close()


def actualizar_sap_empresas() -> dict:
    """
    Actualiza la tabla SAP_EMPRESAS con los datos actuales de SAP HANA.
    SAP HANA es la fuente de verdad.
    - Actualiza empresas existentes con datos de OADM
    - Inserta nuevas empresas
    - Elimina empresas que ya no existen en HANA
    - Preserva el campo SL existente
    """
    # Asegurar que exista la base de datos y tabla
    ensure_database_exists()
    ensure_table_sap_empresas_exists()

    # Obtener empresas actuales de HANA
    empresas_hana = get_empresas_sap()

    mssql_conn = get_mssql_connection()
    mssql_cursor = mssql_conn.cursor()

    resultados = {
        "total_empresas": len(empresas_hana),
        "actualizadas": 0,
        "insertadas": 0,
        "eliminadas": 0,
        "errores": []
    }

    # Eliminar empresas que ya no existen en HANA
    if empresas_hana:
        placeholders = ", ".join(["?" for _ in empresas_hana])
        mssql_cursor.execute(
            f"DELETE FROM SAP_EMPRESAS WHERE Instancia NOT IN ({placeholders})",
            empresas_hana
        )
        resultados["eliminadas"] = mssql_cursor.rowcount
    else:
        mssql_cursor.execute("DELETE FROM SAP_EMPRESAS")
        resultados["eliminadas"] = mssql_cursor.rowcount

    for instancia in empresas_hana:
        try:
            # Verificar si existe versión _PRUEBAS
            tiene_pruebas = schema_exists_in_hana(f"{instancia}_PRUEBAS")

            # Obtener datos de OADM
            oadm = get_oadm_data(instancia)
            print_headr = oadm["PrintHeadr"] if oadm["PrintHeadr"] else instancia

            # Verificar si ya existe en MSSQL
            mssql_cursor.execute(
                "SELECT SL FROM SAP_EMPRESAS WHERE Instancia = ?",
                [instancia]
            )
            row = mssql_cursor.fetchone()

            if row is not None:
                # UPDATE - preservar SL
                mssql_cursor.execute(
                    """
                    UPDATE SAP_EMPRESAS
                    SET Prueba = ?, PrintHeadr = ?, CompnyAddr = ?, TaxIdNum = ?
                    WHERE Instancia = ?
                    """,
                    (
                        1 if tiene_pruebas else 0,
                        print_headr,
                        oadm["CompnyAddr"],
                        oadm["TaxIdNum"],
                        instancia
                    )
                )
                resultados["actualizadas"] += 1
            else:
                # INSERT - SL en 0 por defecto
                mssql_cursor.execute(
                    """
                    INSERT INTO SAP_EMPRESAS (Instancia, SL, Prueba, SLP, PrintHeadr, CompnyAddr, TaxIdNum)
                    VALUES (?, 0, ?, 0, ?, ?, ?)
                    """,
                    (
                        instancia,
                        1 if tiene_pruebas else 0,
                        print_headr,
                        oadm["CompnyAddr"],
                        oadm["TaxIdNum"]
                    )
                )
                resultados["insertadas"] += 1

        except Exception as e:
            resultados["errores"].append({"instancia": instancia, "error": str(e)})

    mssql_conn.commit()
    mssql_cursor.close()
    mssql_conn.close()

    return resultados


def actualizar_sap_proveedores() -> dict:
    """
    Actualiza la tabla SAP_PROVEEDORES con los proveedores de todas las instancias
    que tienen SL habilitado en SAP_EMPRESAS.
    SAP Service Layer es la fuente de verdad.
    - Actualiza proveedores existentes
    - Inserta nuevos proveedores
    - Elimina proveedores que ya no existen en SAP

    Nota: El modo (productivo/pruebas) se controla con la variable global
    configurada mediante set_modo_pruebas() o el endpoint /pruebas/{valor}
    """
    from config import get_instancia_sl, get_modo_pruebas

    # Asegurar que existe la tabla
    ensure_table_sap_proveedores_exists()

    conn = get_mssql_connection()
    cursor = conn.cursor()

    # Obtener instancias con Service Layer habilitado
    instancias = get_instancias_con_service_layer()

    resultados = {
        "modo": "pruebas" if get_modo_pruebas() else "productivo",
        "total_instancias": len(instancias),
        "proveedores_actualizados": 0,
        "proveedores_insertados": 0,
        "proveedores_eliminados": 0,
        "instancias_procesadas": [],
        "errores": []
    }

    # Campos para UPDATE (sin Instancia y CardCode que son la llave)
    campos_update = [
        "CardName", "GroupCode", "FederalTaxID",
        "CreateDate", "CreateTime", "UpdateDate", "UpdateTime",
        "Address", "Block", "ZipCode", "City", "County", "BillToState", "Country",
        "MailAddress", "MailZipCode", "ShipToState", "ShipToDefault",
        "Phone1", "Phone2", "Fax", "Cellular", "EmailAddress", "ContactPerson",
        "PayTermsGrpCode", "PeymentMethodCode", "CreditLimit", "MaxCommitment",
        "DiscountPercent", "PriceListNum", "Currency",
        "DeductibleAtSource", "DeductionPercent", "DeductionValidUntil", "VatGroupLatinAmerica",
        "DefaultBankCode", "DefaultAccount", "BankCountry", "HouseBank", "HouseBankCountry",
        "HouseBankAccount", "HouseBankBranch", "HouseBankIBAN", "IBAN",
        "CreditCardCode", "CreditCardNum", "CreditCardExpiration", "DebitorAccount",
        "CurrentAccountBalance", "OpenDeliveryNotesBalance", "OpenOrdersBalance",
        "OpenChecksBalance", "OpenOpportunities",
        "Valid", "Frozen", "BlockDunning", "BackOrder", "PartialDelivery"
    ]

    # Todos los campos para INSERT
    campos_insert = ["Instancia", "CardCode"] + campos_update

    for instancia in instancias:
        try:
            # Obtener nombre de instancia para Service Layer (agrega _PRUEBAS si está en modo pruebas)
            instancia_sl = get_instancia_sl(instancia)

            # Obtener proveedores de esta instancia desde SAP
            resultado = get_proveedores_sl(instancia_sl)

            if not resultado["success"]:
                resultados["errores"].append({
                    "instancia": instancia,
                    "instancia_sl": instancia_sl,
                    "error": resultado.get("error", "Error desconocido")
                })
                continue

            proveedores = resultado.get("proveedores", [])
            count_actualizados = 0
            count_insertados = 0

            # Obtener CardCodes actuales en SAP para esta instancia
            cardcodes_sap = {prov.get("CardCode") for prov in proveedores}

            # Eliminar proveedores que ya no existen en SAP para esta instancia
            # Nota: SQL Server tiene límite de ~2100 parámetros, así que usamos
            # una tabla temporal para evitar el límite del NOT IN
            if cardcodes_sap:
                # Crear tabla temporal con los CardCodes de SAP
                cursor.execute("""
                    IF OBJECT_ID('tempdb..#CardCodesSAP') IS NOT NULL
                        DROP TABLE #CardCodesSAP
                """)
                cursor.execute("""
                    CREATE TABLE #CardCodesSAP (CardCode NVARCHAR(50) PRIMARY KEY)
                """)

                # Insertar CardCodes en lotes de 1000 para evitar límites
                cardcodes_list = list(cardcodes_sap)
                batch_size = 1000
                for i in range(0, len(cardcodes_list), batch_size):
                    batch = cardcodes_list[i:i + batch_size]
                    placeholders = ", ".join(["(?)" for _ in batch])
                    cursor.execute(
                        f"INSERT INTO #CardCodesSAP (CardCode) VALUES {placeholders}",
                        batch
                    )

                # Eliminar usando JOIN con tabla temporal (más eficiente)
                cursor.execute("""
                    DELETE p FROM SAP_PROVEEDORES p
                    WHERE p.Instancia = ?
                    AND NOT EXISTS (
                        SELECT 1 FROM #CardCodesSAP t WHERE t.CardCode = p.CardCode
                    )
                """, [instancia])
                deleted_count = cursor.rowcount

                # Limpiar tabla temporal
                cursor.execute("DROP TABLE #CardCodesSAP")
            else:
                # Si no hay proveedores en SAP, eliminar todos de esta instancia
                cursor.execute("DELETE FROM SAP_PROVEEDORES WHERE Instancia = ?", [instancia])
                deleted_count = cursor.rowcount

            resultados["proveedores_eliminados"] += deleted_count

            for prov in proveedores:
                # Preparar valores para UPDATE/INSERT
                valores_update = [
                    prov.get("CardName"),
                    prov.get("GroupCode"),
                    prov.get("FederalTaxID"),
                    prov.get("CreateDate"),
                    prov.get("CreateTime"),
                    prov.get("UpdateDate"),
                    prov.get("UpdateTime"),
                    prov.get("Address"),
                    prov.get("Block"),
                    prov.get("ZipCode"),
                    prov.get("City"),
                    prov.get("County"),
                    prov.get("BillToState"),
                    prov.get("Country"),
                    prov.get("MailAddress"),
                    prov.get("MailZipCode"),
                    prov.get("ShipToState"),
                    prov.get("ShipToDefault"),
                    prov.get("Phone1"),
                    prov.get("Phone2"),
                    prov.get("Fax"),
                    prov.get("Cellular"),
                    prov.get("EmailAddress"),
                    prov.get("ContactPerson"),
                    prov.get("PayTermsGrpCode"),
                    prov.get("PeymentMethodCode"),
                    prov.get("CreditLimit"),
                    prov.get("MaxCommitment"),
                    prov.get("DiscountPercent"),
                    prov.get("PriceListNum"),
                    prov.get("Currency"),
                    prov.get("DeductibleAtSource"),
                    prov.get("DeductionPercent"),
                    prov.get("DeductionValidUntil"),
                    prov.get("VatGroupLatinAmerica"),
                    prov.get("DefaultBankCode"),
                    prov.get("DefaultAccount"),
                    prov.get("BankCountry"),
                    prov.get("HouseBank"),
                    prov.get("HouseBankCountry"),
                    prov.get("HouseBankAccount"),
                    prov.get("HouseBankBranch"),
                    prov.get("HouseBankIBAN"),
                    prov.get("IBAN"),
                    prov.get("CreditCardCode"),
                    prov.get("CreditCardNum"),
                    prov.get("CreditCardExpiration"),
                    prov.get("DebitorAccount"),
                    prov.get("CurrentAccountBalance"),
                    prov.get("OpenDeliveryNotesBalance"),
                    prov.get("OpenOrdersBalance"),
                    prov.get("OpenChecksBalance"),
                    prov.get("OpenOpportunities"),
                    prov.get("Valid"),
                    prov.get("Frozen"),
                    prov.get("BlockDunning"),
                    prov.get("BackOrder"),
                    prov.get("PartialDelivery")
                ]

                try:
                    # Verificar si existe
                    cursor.execute(
                        "SELECT 1 FROM SAP_PROVEEDORES WHERE Instancia = ? AND CardCode = ?",
                        [instancia, prov.get("CardCode")]
                    )
                    exists = cursor.fetchone() is not None

                    if exists:
                        # UPDATE
                        update_set = ", ".join([f"{campo} = ?" for campo in campos_update])
                        update_sql = f"""
                            UPDATE SAP_PROVEEDORES
                            SET {update_set}
                            WHERE Instancia = ? AND CardCode = ?
                        """
                        cursor.execute(update_sql, valores_update + [instancia, prov.get("CardCode")])
                        count_actualizados += 1
                    else:
                        # INSERT
                        placeholders = ", ".join(["?" for _ in campos_insert])
                        insert_sql = f"INSERT INTO SAP_PROVEEDORES ({', '.join(campos_insert)}) VALUES ({placeholders})"
                        cursor.execute(insert_sql, [instancia, prov.get("CardCode")] + valores_update)
                        count_insertados += 1

                except Exception as e:
                    resultados["errores"].append({
                        "instancia": instancia,
                        "cardcode": prov.get("CardCode"),
                        "error": str(e)
                    })

            conn.commit()
            resultados["proveedores_actualizados"] += count_actualizados
            resultados["proveedores_insertados"] += count_insertados
            resultados["instancias_procesadas"].append({
                "instancia": instancia,
                "actualizados": count_actualizados,
                "insertados": count_insertados,
                "proveedores": count_actualizados + count_insertados
            })

        except Exception as e:
            resultados["errores"].append({
                "instancia": instancia,
                "error": str(e)
            })

    cursor.close()
    conn.close()

    return resultados


def analizar_inconsistencias_maestro_proveedores() -> dict:
    """
    Analiza inconsistencias en la vista maestro_proveedores y envía reporte por correo.

    Detecta:
    1. Mismo RFC+GroupCode pero diferente CardName
    2. Mismo CardName pero diferente RFC
    3. Mismo RFC con diferentes CardCodes entre instancias
    4. Proveedores con CardName o RFC NULL
    5. Proveedores sin EmailAddress registrado
    """
    settings = get_settings()
    conn = get_mssql_connection()
    cursor = conn.cursor()

    try:
        # Obtener columnas de instancias dinámicamente
        cursor.execute("SELECT DISTINCT Instancia FROM SAP_PROVEEDORES ORDER BY Instancia")
        instancias = [row[0] for row in cursor.fetchall()]

        if not instancias:
            return {"success": False, "error": "No hay instancias en SAP_PROVEEDORES"}

        resultados = {
            "fecha_analisis": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_registros_vista": 0,
            "inconsistencias": {
                "mismo_rfc_diferente_nombre": [],
                "mismo_nombre_diferente_rfc": [],
                "diferentes_cardcodes_por_instancia": [],
                "cardname_null": 0,
                "rfc_null": 0,
                "sin_email": []
            }
        }

        # Total de registros en la vista
        cursor.execute("SELECT COUNT(*) FROM maestro_proveedores")
        resultados["total_registros_vista"] = cursor.fetchone()[0]

        # 1. Mismo RFC+GroupCode pero diferente CardName
        cursor.execute("""
            SELECT
                FederalTaxID,
                GroupCode,
                COUNT(DISTINCT CardName) as nombres_diferentes,
                STRING_AGG(CAST(CardName AS NVARCHAR(MAX)), ' || ') as nombres
            FROM maestro_proveedores
            WHERE FederalTaxID IS NOT NULL
            GROUP BY FederalTaxID, GroupCode
            HAVING COUNT(DISTINCT CardName) > 1
            ORDER BY nombres_diferentes DESC
        """)
        for row in cursor.fetchall():
            resultados["inconsistencias"]["mismo_rfc_diferente_nombre"].append({
                "rfc": row[0],
                "group_code": row[1],
                "nombres_diferentes": row[2],
                "nombres": row[3][:500] if row[3] else ""
            })

        # 2. Mismo CardName pero diferente RFC
        cursor.execute("""
            SELECT
                CardName,
                COUNT(DISTINCT FederalTaxID) as rfcs_diferentes,
                STRING_AGG(FederalTaxID, ', ') as rfcs
            FROM maestro_proveedores
            WHERE CardName IS NOT NULL AND FederalTaxID IS NOT NULL
            GROUP BY CardName
            HAVING COUNT(DISTINCT FederalTaxID) > 1
            ORDER BY rfcs_diferentes DESC
        """)
        for row in cursor.fetchall():
            resultados["inconsistencias"]["mismo_nombre_diferente_rfc"].append({
                "card_name": row[0][:200] if row[0] else "",
                "rfcs_diferentes": row[1],
                "rfcs": row[2][:500] if row[2] else ""
            })

        # 3. Mismo RFC con diferentes CardCodes entre instancias
        # Construir la consulta dinámicamente
        comparaciones = []
        for i in range(len(instancias)):
            for j in range(i + 1, len(instancias)):
                inst1 = instancias[i]
                inst2 = instancias[j]
                comparaciones.append(
                    f"([{inst1}] IS NOT NULL AND [{inst2}] IS NOT NULL AND [{inst1}] <> [{inst2}])"
                )

        if comparaciones:
            where_clause = " OR ".join(comparaciones)
            columnas_instancias = ", ".join([f"[{inst}]" for inst in instancias])

            query = f"""
                SELECT
                    FederalTaxID,
                    CardName,
                    {columnas_instancias}
                FROM maestro_proveedores
                WHERE FederalTaxID IS NOT NULL
                AND ({where_clause})
            """

            cursor.execute(query)
            for row in cursor.fetchall():
                # Contar códigos únicos no nulos
                codigos = [row[i + 2] for i in range(len(instancias)) if row[i + 2] is not None]
                codigos_unicos = len(set(codigos))

                if codigos_unicos > 1:
                    cardcodes_por_inst = {}
                    for i, inst in enumerate(instancias):
                        if row[i + 2] is not None:
                            cardcodes_por_inst[inst] = row[i + 2]

                    resultados["inconsistencias"]["diferentes_cardcodes_por_instancia"].append({
                        "rfc": row[0],
                        "card_name": row[1][:200] if row[1] else "",
                        "codigos_diferentes": codigos_unicos,
                        "total_instancias": len(codigos),
                        "cardcodes": cardcodes_por_inst
                    })

        # 4. CardName NULL
        cursor.execute("SELECT COUNT(*) FROM maestro_proveedores WHERE CardName IS NULL")
        resultados["inconsistencias"]["cardname_null"] = cursor.fetchone()[0]

        # 5. RFC NULL
        cursor.execute("SELECT COUNT(*) FROM maestro_proveedores WHERE FederalTaxID IS NULL")
        resultados["inconsistencias"]["rfc_null"] = cursor.fetchone()[0]

        # 6. Proveedores sin EmailAddress
        # Obtener proveedores de SAP_PROVEEDORES que no tienen email (NULL o vacío)
        cursor.execute("""
            SELECT DISTINCT
                p.CardCode,
                p.CardName,
                p.FederalTaxID,
                p.Instancia
            FROM SAP_PROVEEDORES p
            WHERE (p.EmailAddress IS NULL OR LTRIM(RTRIM(p.EmailAddress)) = '')
            ORDER BY p.CardName, p.Instancia
        """)
        for row in cursor.fetchall():
            resultados["inconsistencias"]["sin_email"].append({
                "card_code": row[0],
                "card_name": row[1][:200] if row[1] else "",
                "rfc": row[2],
                "instancia": row[3]
            })

        # Enviar correo con el reporte
        if settings.EMAIL_SUPERVISOR:
            email_result = enviar_correo_inconsistencias(resultados)
            resultados["email_enviado"] = email_result
        else:
            resultados["email_enviado"] = {"success": False, "error": "EMAIL_SUPERVISOR no configurado"}

        return {
            "success": True,
            "resultados": resultados
        }

    finally:
        cursor.close()
        conn.close()


def enviar_correo_inconsistencias(resultados: dict) -> dict:
    """
    Envía correo con el reporte de inconsistencias del maestro de proveedores.
    Genera un archivo Excel con múltiples hojas para cada tipo de inconsistencia.
    """
    settings = get_settings()

    if not settings.EMAIL_SUPERVISOR:
        return {"success": False, "error": "No hay destinatario configurado"}

    fecha = resultados.get("fecha_analisis", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    subject = f"Reporte de Inconsistencias - Maestro Proveedores - {fecha}"

    inc = resultados["inconsistencias"]

    # Construir cuerpo del mensaje
    body_lines = [
        "REPORTE DE INCONSISTENCIAS - MAESTRO DE PROVEEDORES",
        "=" * 80,
        f"Fecha: {fecha}",
        f"Total registros en vista: {resultados['total_registros_vista']:,}",
        "",
        "RESUMEN DE INCONSISTENCIAS:",
        "-" * 80,
        f"1. Mismo RFC+GroupCode, diferente nombre: {len(inc['mismo_rfc_diferente_nombre']):,} casos",
        f"2. Mismo nombre, diferente RFC: {len(inc['mismo_nombre_diferente_rfc']):,} casos",
        f"3. Diferentes CardCodes entre instancias: {len(inc['diferentes_cardcodes_por_instancia']):,} casos",
        f"4. Proveedores con CardName NULL: {inc['cardname_null']:,}",
        f"5. Proveedores con RFC NULL: {inc['rfc_null']:,}",
        f"6. Proveedores sin EmailAddress: {len(inc['sin_email']):,} casos",
        "",
        "Adjunto encontrará un archivo Excel con el detalle completo de todas las inconsistencias.",
        ""
    ]

    body = "\n".join(body_lines)

    # Generar archivo Excel
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["REPORTE DE INCONSISTENCIAS - MAESTRO DE PROVEEDORES"])
    ws_resumen.append([])
    ws_resumen.append(["Fecha:", fecha])
    ws_resumen.append(["Total registros en vista:", resultados['total_registros_vista']])
    ws_resumen.append([])
    ws_resumen.append(["RESUMEN DE INCONSISTENCIAS:"])
    ws_resumen.append(["Tipo", "Cantidad"])
    ws_resumen.append(["Mismo RFC+GroupCode, diferente nombre", len(inc['mismo_rfc_diferente_nombre'])])
    ws_resumen.append(["Mismo nombre, diferente RFC", len(inc['mismo_nombre_diferente_rfc'])])
    ws_resumen.append(["Diferentes CardCodes entre instancias", len(inc['diferentes_cardcodes_por_instancia'])])
    ws_resumen.append(["Proveedores con CardName NULL", inc['cardname_null']])
    ws_resumen.append(["Proveedores con RFC NULL", inc['rfc_null']])
    ws_resumen.append(["Proveedores sin EmailAddress", len(inc['sin_email'])])

    # Aplicar estilo al encabezado de la tabla
    for cell in ws_resumen[7]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Hoja 2: Mismo RFC diferente nombre
    if inc["mismo_rfc_diferente_nombre"]:
        ws_rfc = wb.create_sheet("RFC - Diferentes Nombres")
        ws_rfc.append(["RFC", "GroupCode", "Nombres Diferentes", "Nombres"])
        for cell in ws_rfc[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for item in inc["mismo_rfc_diferente_nombre"]:
            ws_rfc.append([
                item['rfc'],
                item['group_code'],
                item['nombres_diferentes'],
                item['nombres']
            ])

        # Ajustar anchos de columna
        ws_rfc.column_dimensions['A'].width = 20
        ws_rfc.column_dimensions['B'].width = 15
        ws_rfc.column_dimensions['C'].width = 20
        ws_rfc.column_dimensions['D'].width = 80

    # Hoja 3: Mismo nombre diferente RFC
    if inc["mismo_nombre_diferente_rfc"]:
        ws_nombre = wb.create_sheet("Nombre - Diferentes RFCs")
        ws_nombre.append(["Nombre", "RFCs Diferentes", "RFCs"])
        for cell in ws_nombre[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for item in inc["mismo_nombre_diferente_rfc"]:
            ws_nombre.append([
                item['card_name'],
                item['rfcs_diferentes'],
                item['rfcs']
            ])

        ws_nombre.column_dimensions['A'].width = 60
        ws_nombre.column_dimensions['B'].width = 20
        ws_nombre.column_dimensions['C'].width = 80

    # Hoja 4: Diferentes CardCodes entre instancias
    if inc["diferentes_cardcodes_por_instancia"]:
        ws_cardcodes = wb.create_sheet("Diferentes CardCodes")
        ws_cardcodes.append(["RFC", "Nombre", "Códigos Diferentes", "Total Instancias", "Detalle CardCodes"])
        for cell in ws_cardcodes[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for item in inc["diferentes_cardcodes_por_instancia"]:
            cardcodes_str = " | ".join([f"{inst}: {code}" for inst, code in item['cardcodes'].items()])
            ws_cardcodes.append([
                item['rfc'],
                item['card_name'],
                item['codigos_diferentes'],
                item['total_instancias'],
                cardcodes_str
            ])

        ws_cardcodes.column_dimensions['A'].width = 20
        ws_cardcodes.column_dimensions['B'].width = 60
        ws_cardcodes.column_dimensions['C'].width = 20
        ws_cardcodes.column_dimensions['D'].width = 20
        ws_cardcodes.column_dimensions['E'].width = 80

    # Hoja 5: Proveedores sin EmailAddress
    if inc["sin_email"]:
        ws_email = wb.create_sheet("Sin Email")
        ws_email.append(["CardCode", "Nombre", "RFC", "Instancia"])
        for cell in ws_email[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for item in inc["sin_email"]:
            ws_email.append([
                item['card_code'],
                item['card_name'],
                item['rfc'],
                item['instancia']
            ])

        ws_email.column_dimensions['A'].width = 15
        ws_email.column_dimensions['B'].width = 60
        ws_email.column_dimensions['C'].width = 20
        ws_email.column_dimensions['D'].width = 20

    # Guardar Excel en memoria
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Adjuntar Excel
    attachment = {
        "filename": f"inconsistencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "content": excel_buffer.read()
    }

    return send_email(settings.EMAIL_SUPERVISOR, subject, body, attachment)


def analizar_actividad_proveedores(anos: int = 0) -> dict:
    """
    Analiza la actividad de proveedores y crea/actualiza tablas SAP_PROV_ACTIVOS y SAP_PROV_INACTIVOS.

    El parámetro 'anos' funciona así:
    - 0 = solo año actual
    - 1 = año actual + 1 año anterior (2 años en total)
    - 2 = año actual + 2 años anteriores (3 años en total)

    Metodología:
    1. Consulta OPCH y OPOR en HANA para obtener CardCodes con actividad
    2. Agrupa por CardCode, Instancia
    3. Crea/actualiza tabla SAP_PROV_ACTIVOS con proveedores activos
    4. Cruza con SAP_PROVEEDORES para obtener inactivos
    5. Crea/actualiza tabla SAP_PROV_INACTIVOS
    6. Envía reporte por correo

    Args:
        anos: Años adicionales hacia atrás (0=solo actual, 1=actual+1 anterior, etc.)

    Returns:
        dict con resultados del análisis
    """
    settings = get_settings()
    conn_mssql = get_mssql_connection()
    cursor_mssql = conn_mssql.cursor()

    try:
        # Obtener todas las instancias SAP
        cursor_mssql.execute("SELECT Instancia FROM SAP_EMPRESAS ORDER BY Instancia")
        instancias = [row[0] for row in cursor_mssql.fetchall()]

        if not instancias:
            return {"success": False, "error": "No hay instancias en SAP_EMPRESAS"}

        # Eliminar tablas si existen
        cursor_mssql.execute("DROP TABLE IF EXISTS SAP_PROV_ACTIVOS")
        cursor_mssql.execute("DROP TABLE IF EXISTS SAP_PROV_INACTIVOS")
        conn_mssql.commit()

        # Crear tabla SAP_PROV_ACTIVOS
        cursor_mssql.execute("""
            CREATE TABLE SAP_PROV_ACTIVOS (
                Instancia NVARCHAR(50) NOT NULL,
                CardCode NVARCHAR(50) NOT NULL,
                CardName NVARCHAR(200),
                FederalTaxID NVARCHAR(50),
                GroupCode INT,
                TotalDocumentos INT,
                UltimaFecha DATE,
                FechaAnalisis DATETIME,
                PRIMARY KEY (Instancia, CardCode)
            )
        """)

        # Crear tabla SAP_PROV_INACTIVOS
        cursor_mssql.execute("""
            CREATE TABLE SAP_PROV_INACTIVOS (
                Instancia NVARCHAR(50) NOT NULL,
                CardCode NVARCHAR(50) NOT NULL,
                CardName NVARCHAR(200),
                FederalTaxID NVARCHAR(50),
                GroupCode INT,
                FechaAnalisis DATETIME,
                PRIMARY KEY (Instancia, CardCode)
            )
        """)
        conn_mssql.commit()

        resultados = {
            "fecha_analisis": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "anos_analizados": anos,
            "total_activos": 0,
            "total_inactivos": 0,
            "instancias_procesadas": []
        }

        # Conectar a HANA
        conn_hana = get_hana_connection()
        cursor_hana = conn_hana.cursor()

        # Procesar cada instancia
        for instancia in instancias:
            try:
                # Consultar CardCodes con actividad en HANA
                # Obtener de OPCH y OPOR los CardCodes activos
                query = f"""
                    SELECT
                        "CardCode",
                        COUNT(*) as total_docs,
                        MAX("DocDate") as ultima_fecha
                    FROM (
                        SELECT "CardCode", "DocDate"
                        FROM "{instancia}".OPCH
                        WHERE YEAR("DocDate") >= YEAR(CURRENT_DATE) - {anos}
                        UNION ALL
                        SELECT "CardCode", "DocDate"
                        FROM "{instancia}".OPOR
                        WHERE YEAR("DocDate") >= YEAR(CURRENT_DATE) - {anos}
                    )
                    GROUP BY "CardCode"
                """

                cursor_hana.execute(query)
                cardcodes_activos = cursor_hana.fetchall()

                activos_instancia = 0
                inactivos_instancia = 0

                # Insertar proveedores activos en SAP_PROV_ACTIVOS
                for card_code, total_docs, ultima_fecha in cardcodes_activos:
                    # Obtener información del proveedor de SAP_PROVEEDORES
                    cursor_mssql.execute("""
                        SELECT CardName, FederalTaxID, GroupCode
                        FROM SAP_PROVEEDORES
                        WHERE Instancia = ? AND CardCode = ?
                    """, [instancia, card_code])

                    result = cursor_mssql.fetchone()
                    if result:
                        card_name, rfc, group_code = result

                        cursor_mssql.execute("""
                            INSERT INTO SAP_PROV_ACTIVOS
                            (Instancia, CardCode, CardName, FederalTaxID, GroupCode, TotalDocumentos, UltimaFecha, FechaAnalisis)
                            VALUES (?, ?, ?, ?, ?, ?, ?, GETDATE())
                        """, [instancia, card_code, card_name, rfc, group_code, total_docs, ultima_fecha])

                        activos_instancia += 1

                # Insertar proveedores inactivos (los que están en SAP_PROVEEDORES pero NO en SAP_PROV_ACTIVOS)
                cursor_mssql.execute("""
                    INSERT INTO SAP_PROV_INACTIVOS (Instancia, CardCode, CardName, FederalTaxID, GroupCode, FechaAnalisis)
                    SELECT
                        p.Instancia,
                        p.CardCode,
                        p.CardName,
                        p.FederalTaxID,
                        p.GroupCode,
                        GETDATE()
                    FROM SAP_PROVEEDORES p
                    WHERE p.Instancia = ?
                    AND NOT EXISTS (
                        SELECT 1 FROM SAP_PROV_ACTIVOS a
                        WHERE a.Instancia = p.Instancia AND a.CardCode = p.CardCode
                    )
                """, [instancia])

                inactivos_instancia = cursor_mssql.rowcount
                conn_mssql.commit()

                resultados["total_activos"] += activos_instancia
                resultados["total_inactivos"] += inactivos_instancia
                resultados["instancias_procesadas"].append({
                    "instancia": instancia,
                    "activos": activos_instancia,
                    "inactivos": inactivos_instancia
                })

            except Exception as e:
                resultados["instancias_procesadas"].append({
                    "instancia": instancia,
                    "error": str(e)
                })

        cursor_hana.close()
        conn_hana.close()

        # Enviar correo con el reporte
        if settings.EMAIL_SUPERVISOR:
            email_result = enviar_correo_actividad_proveedores(anos)
            resultados["email_enviado"] = email_result
        else:
            resultados["email_enviado"] = {"success": False, "error": "EMAIL_SUPERVISOR no configurado"}

        return {
            "success": True,
            "resultados": resultados
        }

    finally:
        cursor_mssql.close()
        conn_mssql.close()


def enviar_correo_actividad_proveedores(anos: int) -> dict:
    """
    Envía correo con el reporte de actividad de proveedores.
    Genera un archivo Excel con dos hojas: Activos e Inactivos.
    Lee los datos de las tablas SAP_PROV_ACTIVOS y SAP_PROV_INACTIVOS.
    """
    settings = get_settings()

    if not settings.EMAIL_SUPERVISOR:
        return {"success": False, "error": "No hay destinatario configurado"}

    conn = None
    try:
        conn = get_mssql_connection()
        cursor = conn.cursor()

        # Consultar proveedores activos agrupados por CardName, FederalTaxID, GroupCode
        query_activos = """
            SELECT
                CardName,
                FederalTaxID,
                STRING_AGG(CardCode, ', ') WITHIN GROUP (ORDER BY Instancia) as cardcodes,
                MAX(GroupCode) as GroupCode,
                SUM(TotalDocumentos) as total_documentos,
                STRING_AGG(Instancia, ', ') WITHIN GROUP (ORDER BY Instancia) as instancias,
                MAX(UltimaFecha) as ultima_fecha
            FROM SAP_PROV_ACTIVOS
            GROUP BY CardName, FederalTaxID
            ORDER BY CardName
        """
        cursor.execute(query_activos)
        activos = cursor.fetchall()

        # Consultar proveedores inactivos agrupados por CardName, FederalTaxID, GroupCode
        query_inactivos = """
            SELECT DISTINCT
                CardName,
                FederalTaxID,
                STRING_AGG(CardCode, ', ') WITHIN GROUP (ORDER BY Instancia) as cardcodes,
                MAX(GroupCode) as GroupCode
            FROM SAP_PROV_INACTIVOS
            GROUP BY CardName, FederalTaxID
            ORDER BY CardName
        """
        cursor.execute(query_inactivos)
        inactivos = cursor.fetchall()

        # Contar total de proveedores únicos en SAP_PROVEEDORES
        cursor.execute("SELECT COUNT(DISTINCT CardName + ISNULL(FederalTaxID, '')) FROM SAP_PROVEEDORES")
        total_proveedores = cursor.fetchone()[0]

        cursor.close()

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Calcular período de años involucrados
        ano_actual = datetime.now().year
        ano_inicio = ano_actual - anos
        if anos == 0:
            periodo = f"{ano_actual}"
        elif anos == 1:
            periodo = f"{ano_inicio}-{ano_actual}"
        else:
            periodo = f"{ano_inicio}-{ano_actual}"

        subject = f"Actividad de Proveedores - Período {periodo} - {fecha}"

        total_activos = len(activos)
        total_inactivos = len(inactivos)

        # Construir cuerpo del mensaje
        body_lines = [
            "REPORTE DE ACTIVIDAD DE PROVEEDORES",
            "=" * 80,
            f"Fecha: {fecha}",
            f"Período analizado: {periodo}",
            f"Total proveedores únicos: {total_proveedores:,}",
            "",
            "RESUMEN:",
            "-" * 80,
            f"Proveedores ACTIVOS: {total_activos:,} ({total_activos/total_proveedores*100:.1f}%)" if total_proveedores > 0 else f"Proveedores ACTIVOS: {total_activos:,}",
            f"Proveedores INACTIVOS: {total_inactivos:,} ({total_inactivos/total_proveedores*100:.1f}%)" if total_proveedores > 0 else f"Proveedores INACTIVOS: {total_inactivos:,}",
            "",
            "Adjunto encontrará un archivo Excel con el detalle completo de la actividad.",
            ""
        ]

        body = "\n".join(body_lines)

        # Generar archivo Excel
        wb = Workbook()

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Hoja 1: Activos
        ws_activos = wb.active
        ws_activos.title = "Activos"
        ws_activos.append(["CardName", "FederalTaxID", "CardCode", "GroupCode", "Total Documentos", "Instancias con Actividad", "Última Fecha"])

        for cell in ws_activos[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in activos:
            ws_activos.append([
                row[0],  # CardName
                row[1],  # FederalTaxID
                row[2],  # CardCode
                row[3],  # GroupCode
                row[4],  # total_documentos
                row[5],  # instancias
                row[6].strftime("%Y-%m-%d") if row[6] else ""  # ultima_fecha
            ])

        ws_activos.column_dimensions['A'].width = 60
        ws_activos.column_dimensions['B'].width = 20
        ws_activos.column_dimensions['C'].width = 20
        ws_activos.column_dimensions['D'].width = 15
        ws_activos.column_dimensions['E'].width = 20
        ws_activos.column_dimensions['F'].width = 40
        ws_activos.column_dimensions['G'].width = 15

        # Hoja 2: Inactivos
        ws_inactivos = wb.create_sheet("Inactivos")
        ws_inactivos.append(["CardName", "FederalTaxID", "CardCode", "GroupCode"])

        for cell in ws_inactivos[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in inactivos:
            ws_inactivos.append([
                row[0],  # CardName
                row[1],  # FederalTaxID
                row[2],  # CardCode
                row[3]   # GroupCode
            ])

        ws_inactivos.column_dimensions['A'].width = 60
        ws_inactivos.column_dimensions['B'].width = 20
        ws_inactivos.column_dimensions['C'].width = 20
        ws_inactivos.column_dimensions['D'].width = 15

        # Guardar Excel en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Adjuntar Excel
        attachment = {
            "filename": f"Actividad_Proveedores_Periodo_{periodo.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "content": excel_buffer.read()
        }

        return send_email(settings.EMAIL_SUPERVISOR, subject, body, attachment)

    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        if conn:
            conn.close()


# Alias para compatibilidad con código existente
def poblar_sap_proveedores() -> dict:
    """Alias de actualizar_sap_proveedores() para compatibilidad."""
    return actualizar_sap_proveedores()
